import requests
import json
from openpyxl import Workbook
import itertools
import datetime
from random import randint



def epoch_to_date(epoch_time):
        time_stamp = epoch_time / 1000
        date_time = datetime.datetime.fromtimestamp(time_stamp).strftime('%Y-%m-%d %H:%M:%S')
        return date_time

def download_stat():
        
    url = f"https://open-api.coinglass.com/public/v2/long_short_history?time_type=h1&symbol=BTC"

    headers = {
        "accept": "application/json",
        "coinglassSecret": "1ab873cda0e44304a955ff179fae87d2"
    }

    response = requests.get(url, headers=headers)
    todos = json.loads(response.text)
    data = todos['data']
    symbol = data['symbol']
    longRateList = data['longRateList']
    shortsRateList = data['shortsRateList']
    priceList = data['priceList']
    dateList = data['dateList']

    wb = Workbook()
    ws = wb.active      
    ws['A1'] = 'Дата'
    ws['B1'] = 'Лонг'
    ws['C1'] = 'Шорт'
    ws['D1'] = 'Цена'

    row = 2

    

    for (g, i, o, p) in itertools.zip_longest(dateList, longRateList, shortsRateList, priceList):
        ws[row][0].value = epoch_to_date(g)
        ws[row][1].value = i
        ws[row][2].value = o
        ws[row][3].value = p
        row += 1 

    wb.save(f'{datetime.datetime.now().date()}_days_id{randint(0, 3000)}.xlsx')

